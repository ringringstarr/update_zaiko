import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime
import io
import re

# --- 1. アプリの画面設定 ---
st.set_page_config(page_title="在庫表 自動更新ツール", page_icon="♨️")
st.title("♨️ 在庫表 自動更新ツール")

st.write("最新の『年度在庫速度表.xlsx』を下の枠にドラッグ＆ドロップしてください。")

# --- 2. ファイルアップローダー（常に一番上） ---
uploaded_file = st.file_uploader("エクセルファイルを選択", type=['xlsx'])

# 処理結果を一時保存する変数
processed_file = None
save_filename = ""

# --- 3. 処理ロジック ---
if uploaded_file is not None:
    with st.spinner('在庫データを比較・更新しています...'):
        try:
            wb = openpyxl.load_workbook(uploaded_file)
            ws_zenjitsu = wb['新前日']
            ws_toujitsu = wb['新当日']
            ws_zaiko = wb['半露在庫']

            TARGET_ROOMS = [
                '11 半露', '12 半露', '13 半露', '01 半露', 
                '02 半露', '03 半露', '露天風呂付客室', '源泉内風呂付' 
            ]

            COLORS = {
                '0': '1E90FF', '1': 'FFDAE0', '2': '7FFFD4', '3': 'D8BFD8',
                '4': 'B0E0E6', '5': 'FFDEAD', '6': 'E8997A', '7': '4A84B6',
                '8': '6B9027', '9': '6A5ACD', '10': '708090', '11': '00FFFF',
                '12': 'B0C4DE', '13': '7CFC00'
            }

            def get_inventory(val):
                val_str = str(val).strip()
                if val_str in ['1', '1.0']:
                    return 1
                elif val_str in ['0', '0.0', 'None', '']:
                    return 0
                return None

            for row in range(2, ws_zaiko.max_row + 1):
                room_name = ws_zaiko.cell(row=row, column=1).value
                
                if room_name in TARGET_ROOMS:
                    for col in range(2, ws_zaiko.max_column + 1):
                        cell_zaiko = ws_zaiko.cell(row=row, column=col)
                        current_val_str = str(cell_zaiko.value).strip()

                        val_zen = get_inventory(ws_zenjitsu.cell(row=row, column=col).value)
                        val_tou = get_inventory(ws_toujitsu.cell(row=row, column=col).value)
                        
                        if val_zen is None or val_tou is None:
                            continue

                        # --- ずれチェック ---
                        is_excel_available = (current_val_str in COLORS.keys()) or (current_val_str == 'キャンセル')
                        is_excel_booked = current_val_str in ['-', '売']
                        
                        mismatch = False
                        if val_zen == 0 and is_excel_booked:
                            mismatch = True
                        elif val_zen == 1 and is_excel_available:
                            mismatch = True
                        
                        if mismatch:
                            cell_zaiko.value = '要確認'
                            cell_zaiko.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            cell_zaiko.font = Font(name='ＭＳ Ｐゴシック', bold=True, color='FF0000')
                            cell_zaiko.alignment = Alignment(horizontal='center')
                            continue

                        # --- 正常処理 ---
                        if cell_zaiko.value == '売':
                            cell_zaiko.value = '-'

                        if val_zen == 0 and val_tou == 1:
                            cell_zaiko.value = '売'
                        elif val_zen == 1 and val_tou == 0:
                            cell_zaiko.value = 'キャンセル'
                        
                        # --- 書式設定（フォント出し分け） ---
                        val_str = str(cell_zaiko.value).strip()

                        if val_str == '-':
                            cell_zaiko.alignment = Alignment(horizontal='right')
                            cell_zaiko.font = Font(name='Arial', bold=False)
                            cell_zaiko.fill = PatternFill(fill_type=None)
                        
                        elif val_str == '売':
                            cell_zaiko.alignment = Alignment(horizontal='center')
                            cell_zaiko.font = Font(name='ＭＳ Ｐゴシック', bold=True)
                            cell_zaiko.fill = PatternFill(fill_type=None)
                        
                        elif val_str == 'キャンセル':
                            cell_zaiko.alignment = Alignment(horizontal='center')
                            cell_zaiko.font = Font(name='ＭＳ Ｐゴシック', bold=True, color='FF0000') 
                            cell_zaiko.fill = PatternFill(fill_type=None)
                        
                        elif val_str in COLORS:
                            fill_color = COLORS[val_str]
                            cell_zaiko.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                            cell_zaiko.alignment = Alignment(horizontal='center')
                            cell_zaiko.font = Font(name='ＭＳ Ｐゴシック', bold=False, color='000000')

            # --- メモリ上に保存 ---
            today_str = datetime.today().strftime('%Y.%m.%d')
            clean_name = re.sub(r'^\d{4}\.\d{2}\.\d{2}\s*', '', uploaded_file.name)
            save_filename = f"{today_str} {clean_name}"

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 成功したら変数を書き換える
            processed_file = output

        except Exception as e:
            st.error(f"エラーが発生しました: シート名やフォーマットが正しいか確認してください。（エラー詳細：{e}）")

    # --- ダウンロードボタンの表示（処理が成功した場合のみ、ここに必ず出る） ---
    if processed_file is not None:
        st.success('処理が完了しました！下のボタンから新しい在庫表をダウンロードしてください。')
        st.download_button(
            label="📥 更新版をダウンロードする",
            data=processed_file,
            file_name=save_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.write("---")

# --- 4. マニュアル（仕様書）の表示部分（絶対に一番下になるように配置） ---
with st.expander("使用方法", expanded=False):
    st.markdown("""
    **新滝の温泉付き客室の在庫の増減表を自動で作成するツールです。**
    正確なデータに基づいた在庫増減の確認ができます。

    ### ① 日々の在庫速度表を作成する
    以下のシェアフォルダ内でいつも通り、速度表の作成を開始します。
    `\\\\192.168.1.222\\share\\〇マーケティング\\●202〇年度在庫速度表`
    
    1. 支配人君から「日別部屋タイプ集計」をダウンロードし、「新当日」にコピペします。
    2. 在庫は「新前日」,「新当日」シートの増減を基に計算します。
    3. 「半露在庫」の不要な日程を削除してください。（前日分など）
    
    ⇒ **下準備はこれで完成です。**

    ### ② 在庫表をドロップしてください
    1. 上のアップロード枠に、下準備が終わった在庫表をドロップしてください。
    2. 在庫の増減を反映した新しいエクセルファイルをダウンロードできます。
    3. その後、再度また `\\\\192.168.1.222\\share\\〇マーケティング\\●202〇年度在庫速度表` にファイルを戻し、上書きしてください。

    ---
    ### ⚙️ 機能とルール
    * **在庫の増減判定ルール**
      「新前日」と「新当日」の数字（0＝空室、1＝満室）を比較し、以下の判定を行います。
      * 売れた場合はセルに「**売**」と入力されます。
      * キャンセルが出た場合はセルに「**キャンセル**」と入力されます。
        ⚠️ **キャンセルの箇所は手動でランク数字を入力してください。**
      * 前日の「売」は自動で「 **-** 」に変換されます。
    * **エラーチェック**
      誤入力が発生している場合、「**要確認**」のアラートがでます。
      ⚠️ **手動で修正してください。**
    """)

